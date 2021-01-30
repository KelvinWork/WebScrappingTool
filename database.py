from openpyxl import Workbook
from openpyxl import load_workbook

import datetime


def createWorkBook():
    print("- -writeWorkBook Called")
    wb = Workbook()
    dest_filename = 'data_extraction_test.xlsx'
    wb.save(filename=dest_filename)


def writeWorkBook(price, alphabet):
    print("- -writeWorkBook Called")
    wb = load_workbook(filename="data_extraction_test.xlsx")
    ws1 = wb.active
    ws1.title = "Sheet"  # sheet name

    cellBlock = alphabet
    cellNumber = 1
    #print("This is the Block letter {}".format(cellBlock))
    #print(cellBlock)
    #print("this is len {}".format(len(cellBlock)))
    itemsInCell = ws1["{}{}".format(cellBlock, cellNumber)].value
    #print(type(itemsInCell))
    #print(cellBlock)
    #print(cellNumber)

    while (itemsInCell != None):

        itemsInCell = ws1['{}{}'.format(cellBlock, cellNumber)].value
        #print("this is in itemInCell {}".format(itemsInCell))

        cellNumber += 1
        #print(cellNumber)
        #print(cellBlock)

        if (itemsInCell == None):
            ws1['{}{}'.format(cellBlock, (cellNumber - 1))] = price
            break

    #print("write function has been called")
    #print(cellNumber)
    wb.save(filename="data_extraction_test.xlsx")

def writeDateExtracted():
    print("- -writeDateExtract Called")
    wb = load_workbook(filename="data_extraction_test.xlsx")
    ws1 = wb.active
    ws1.title = "Sheet"  # sheet name


    todayDate = datetime.datetime.now().date()
    ws1['A1'] = "Date"
    dateBlock = "A"
    dateCellNumber = 1

    #writing of the date

    print(todayDate)
    dateInCell = ws1['{}{}'.format(dateBlock, dateCellNumber)].value

    while(dateInCell != None):
        dateInCell = ws1['{}{}'.format(dateBlock, dateCellNumber)].value
        dateCellNumber += 1

        if(dateInCell == None):
            ws1['{}{}'.format(dateBlock, (dateCellNumber - 1))] = todayDate
            break

    wb.save(filename="data_extraction_test.xlsx")



def readWorkBook(cells):
    print("- -readWorkBook Called")
    wb = load_workbook(filename="data_extraction_test.xlsx", )
    sheet_ranges = wb['Sheet']
    print(sheet_ranges[cells].value)
    theReturnString = "Get Good"
    return theReturnString


def writeHeaderProduct(productName, alphabet):
    print("- -writeHeaderProduct Called")
    wb = load_workbook(filename="data_extraction_test.xlsx")
    ws1 = wb.active
    ws1.title = "Sheet"
    ws1['{}1'.format(alphabet)] = productName
    wb.save(filename="data_extraction_test.xlsx")
    #print("test function called")




#readWorkBook("C6")

#writeWorkBook("89.90")









